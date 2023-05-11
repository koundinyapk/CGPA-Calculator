from django.shortcuts import render,redirect
from openpyxl import load_workbook
import tempfile
from django.http import HttpRequest,HttpResponse,FileResponse

# Create your views here.
def home(request):
    return render(request,'sgpa.html')


def sgpa(request):
    if request.method=="POST":
        subGrade=request.FILES["subGrade"]
        blankSgpaFile=request.FILES["blankSgpaFile"]
        wb_subGrade=load_workbook(subGrade)
        wb_blankSgpaFile=load_workbook(blankSgpaFile)
        ws_blankSgpaFile=wb_blankSgpaFile.active
        noOfSub=len(wb_subGrade.worksheets)
        shift=0
        for ws in wb_subGrade.worksheets:
            last_row=ws.max_row
            for row in range(3,last_row+1):
                grade_src=ws.cell(row=row,column=9).value
                points_src=ws.cell(row=row,column=10).value
                ws_blankSgpaFile.cell(row=row+1,column=5+shift).value=grade_src
                ws_blankSgpaFile.cell(row=row+1,column=6+shift).value=points_src
            shift=shift+3 
        wb_blankSgpaFile.save(blankSgpaFile)

        ws_blankSgpaFile=wb_blankSgpaFile.active
        last_row=ws_blankSgpaFile.max_row
        
        for row in range(4,last_row+1):
            n=noOfSub
            numerator=0
            denominator=0
            shift=0
            while(n>0):
                credits=ws_blankSgpaFile.cell(row=row,column=4+shift).value
                points=ws_blankSgpaFile.cell(row=row,column=6+shift).value
                '''print("--------------"+str(credits)+"-----------------")
                print("------------------"+str(points)+"---------------")
                n-=1
                shift+=3'''
                numerator+=(credits*points)
                denominator+=credits
                n-=1
                shift+=3
            gpa=round(numerator/denominator,2)
            ws_blankSgpaFile.cell(row=row,column=(4+(noOfSub*3))).value=gpa
        
        wb_blankSgpaFile.save(blankSgpaFile)

        #return HttpResponse("Success")
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            wb_blankSgpaFile.save(tmp.name)

        response = FileResponse(open(tmp.name, 'rb'), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="SgpaResult.xlsx"'
        return response
    return HttpResponse("Unsuccessful")
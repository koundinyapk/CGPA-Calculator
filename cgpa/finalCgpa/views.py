from django.shortcuts import render,redirect
from openpyxl import load_workbook
import tempfile
from django.http import HttpRequest,HttpResponse,FileResponse

# Create your views here.
def home(request):
    return render(request,'cgpa.html')

def cgpa_cal(request):
    if request.method=='POST':
        sgpa_file=request.FILES["sgpa_file"]
        cgpa_header=request.FILES["cgpa_header"]
        wb_sgpaFile=load_workbook(sgpa_file)
        wb_cgpaHeader=load_workbook(cgpa_header)
        semester=request.POST.get('semester')
        ws_sgpa=wb_sgpaFile.active
        ws_cgpa=wb_cgpaHeader.active
        sgpa_col=ws_sgpa.max_column
        current_sem_col=3+(2*int(semester))
        last_row=ws_sgpa.max_row
        for row in range(4,last_row+1):
            ws_cgpa.cell(row=row+1,column=current_sem_col).value=ws_sgpa.cell(row=row,column=sgpa_col).value
        wb_cgpaHeader.save(cgpa_header)

        cgpa_col=ws_cgpa.max_column
        cgpa_last_row=ws_cgpa.max_row
        for row in range(5,cgpa_last_row+1):
            numerator=0
            denominator=0
            shift=0
            n=int(semester)
            while(n>0):
                credits=ws_cgpa.cell(row=row,column=4+shift).value
                sgpa=ws_cgpa.cell(row=row,column=5+shift).value
                numerator+=credits*sgpa
                denominator+=credits
                shift+=2
                n-=1
            cgpa=round(numerator/denominator,2)
            ws_cgpa.cell(row=row,column=cgpa_col).value=cgpa
        wb_cgpaHeader.save(cgpa_header)
        #return HttpResponse("Success")
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            wb_cgpaHeader.save(tmp.name)

        response = FileResponse(open(tmp.name, 'rb'), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="CgpaResult.xlsx"'
        return response
    return HttpResponse("Unsuccessful")

from django.shortcuts import render,redirect
from openpyxl import load_workbook
#from openpyxl.utils import sort
import tempfile
from django.http import HttpResponse, FileResponse
#import pandas as pd

# Create your views here.

def home(request):
    return render(request,'eachSub.html')


def sub_wise(request):
    if request.method=='POST':
        #print(request.FILES)
        excel_file=request.FILES["excel_file"]
        wb=load_workbook(excel_file)
        for ws in wb.worksheets:
            #ws=wb.active
            last_row=ws.max_row
            for r in range(3,last_row+1):
                cell=ws.cell(row=r,column=8)
                cell.value=ws.cell(row=r,column=4).value+ws.cell(row=r,column=5).value+ws.cell(row=r,column=6).value+ws.cell(row=r,column=7).value

            wb.save(excel_file)

            #ws=wb.active
            rows=list(ws.rows)
            sorted_rows=sorted(rows[2:last_row+1],key=lambda row: row[7].value,reverse=True)
            ws.delete_rows(3,last_row-2)
            for r,row in enumerate(sorted_rows,start=3):
                for c,cell in enumerate(row,start=1):
                    ws.cell(row=r,column=c).value=cell.value
            wb.save(excel_file)
            

            for r in range(3,last_row+1):
                if ws.cell(row=r,column=8).value<35 or ws.cell(row=r,column=7).value<17.5:
                    ws.cell(row=r,column=9).value='F'
                    ws.cell(row=r,column=10).value=0
                else:
                    ex=int(round(0.05*(last_row-3+1),0))
                    a=int(round(0.15*(last_row-3+1),0))
                    b=int(round(0.3*(last_row-3+1),0))
                    c=int(round(0.5*(last_row-3+1),0))
                    d=int(round(0.75*(last_row-3+1),0))
                    p=int(round((last_row-3+1),0))

                    for r in range(1,ex+1):
                        if ws.cell(row=r+2,column=9).value!='F':
                            ws.cell(row=r+2,column=9).value='Ex'
                            ws.cell(row=r+2,column=10).value=10
                    for r in range(ex+1,a+1):
                        if ws.cell(row=r+2,column=9).value!='F':
                            ws.cell(row=r+2,column=9).value='A'
                            ws.cell(row=r+2,column=10).value=9
                    for r in range(a+1,b+1):
                        if ws.cell(row=r+2,column=9).value!='F':
                            ws.cell(row=r+2,column=9).value='B'
                            ws.cell(row=r+2,column=10).value=8
                    for r in range(b+1,c+1):
                        if ws.cell(row=r+2,column=9).value!='F':
                            ws.cell(row=r+2,column=9).value='C'
                            ws.cell(row=r+2,column=10).value=7
                    for r in range(c+1,d+1):
                        if ws.cell(row=r+2,column=9).value!='F':
                            ws.cell(row=r+2,column=9).value='D'
                            ws.cell(row=r+2,column=10).value=6
                    for r in range(d+1,p+1):
                        if ws.cell(row=r+2,column=9).value!='F':
                            ws.cell(row=r+2,column=9).value='P'
                            ws.cell(row=r+2,column=10).value=5

            wb.save(excel_file)

            #ws=wb.active
            rows=list(ws.rows)
            sorted_rows=sorted(rows[2:last_row+1],key=lambda row: row[0].value)
            ws.delete_rows(3,last_row-2)
            for r,row in enumerate(sorted_rows,start=3):
                for c,cell in enumerate(row,start=1):
                    ws.cell(row=r,column=c).value=cell.value
            wb.save(excel_file)
    
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            wb.save(tmp.name)

        response = FileResponse(open(tmp.name, 'rb'), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="result.xlsx"'
        return response
        #return HttpResponse("Success")
    return HttpResponse("Unsuccessful")